"""Document storage and text extraction for uploaded investigation files.

Uploaded binaries are written to ``data/documents/{investigation_id}/`` and
only metadata is kept in the investigation record. Text extraction is best-effort
and degrades gracefully when an optional dependency or OCR engine is missing.
"""

from __future__ import annotations

import csv
import hashlib
import logging
import re
import threading
from datetime import datetime, timezone
from pathlib import Path
from typing import Optional

logger = logging.getLogger("opencredit.documents")

_EXTRACTED_TEXT_LIMIT = 20_000


def _safe_filename(name: str) -> str:
    """Return a filesystem-safe version of an uploaded filename."""
    base = name.split("/")[-1].split("\\")[-1].strip()
    if not base:
        base = "upload"
    base = re.sub(r"[^\w.\-]", "_", base)
    if base.startswith("."):
        base = "upload" + base
    return base[:120]


def _ext(path: str | Path) -> str:
    return Path(path).suffix.lower().lstrip(".")


def _now() -> str:
    return datetime.now(timezone.utc).isoformat()


class DocumentStorage:
    """Saves uploaded files under ``directory/{investigation_id}/``."""

    def __init__(self, directory: str) -> None:
        self._directory = Path(directory)
        self._directory.mkdir(parents=True, exist_ok=True)
        self._lock = threading.Lock()

    def _investigation_dir(self, investigation_id: str) -> Path:
        path = self._directory / investigation_id
        path.mkdir(parents=True, exist_ok=True)
        return path

    def save(self, investigation_id: str, upload_file) -> dict:
        """Persist an uploaded file and return its metadata."""
        filename = _safe_filename(upload_file.filename or "upload")
        file_bytes = upload_file.file.read()
        sha256 = hashlib.sha256(file_bytes).hexdigest()

        dest = self._investigation_dir(investigation_id) / filename
        original_dest = dest
        stem = dest.stem
        suffix = dest.suffix
        counter = 1
        while dest.exists():
            dest = original_dest.with_name(f"{stem}_{counter}{suffix}")
            counter += 1

        with self._lock:
            dest.write_bytes(file_bytes)

        return {
            "filename": filename,
            "stored_name": dest.name,
            "stored_path": str(dest),
            "content_type": upload_file.content_type or "application/octet-stream",
            "size": len(file_bytes),
            "sha256": sha256,
            "extracted": False,
            "extracted_text": None,
            "extracted_at": None,
        }

    def extract(self, metadata: dict) -> Optional[str]:
        """Extract text from a stored document and update its metadata."""
        path = Path(metadata["stored_path"])
        if not path.exists():
            return None
        text = extract_text(path, metadata.get("content_type"))
        metadata["extracted"] = True
        metadata["extracted_text"] = text
        metadata["extracted_at"] = _now()
        return text


def extract_text(path: str | Path, content_type: Optional[str] = None) -> str:
    """Dispatch extraction based on file extension and content type."""
    path = Path(path)
    ext = _ext(path)
    ct = (content_type or "").lower()

    try:
        if ext == "pdf" or ct == "application/pdf":
            return _extract_pdf(path)
        if ext == "docx" or ct == (
            "application/vnd.openxmlformats-officedocument.wordprocessingml.document"
        ):
            return _extract_docx(path)
        if ext == "xlsx" or ct == (
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        ):
            return _extract_xlsx(path)
        if ext == "csv" or ct == "text/csv":
            return _extract_csv(path)
        if ext == "txt" or ct.startswith("text/"):
            return _extract_txt(path)
        if ext in {"jpg", "jpeg", "png"} or ct.startswith("image/"):
            return _extract_image(path)
    except Exception as exc:
        logger.warning("Document extraction failed for %s: %s", path, exc)

    logger.info("No extractor available for %s (content-type: %s)", path, ct)
    return ""


def _extract_pdf(path: Path) -> str:
    try:
        from pypdf import PdfReader
    except ImportError:
        logger.warning("pypdf is not installed; PDF text extraction skipped.")
        return ""

    reader = PdfReader(str(path))
    parts = []
    for page in reader.pages[:10]:
        try:
            text = page.extract_text()
        except Exception:
            text = ""
        if text:
            parts.append(text)
    return "\n\n".join(parts)[:_EXTRACTED_TEXT_LIMIT]


def _extract_docx(path: Path) -> str:
    try:
        from docx import Document
    except ImportError:
        logger.warning("python-docx is not installed; DOCX extraction skipped.")
        return ""

    doc = Document(str(path))
    paragraphs = [p.text for p in doc.paragraphs if p.text]
    return "\n".join(paragraphs)[:_EXTRACTED_TEXT_LIMIT]


def _extract_xlsx(path: Path) -> str:
    try:
        from openpyxl import load_workbook
    except ImportError:
        logger.warning("openpyxl is not installed; XLSX extraction skipped.")
        return ""

    wb = load_workbook(str(path), data_only=True, read_only=True)
    parts = []
    for sheet in wb.worksheets[:3]:
        rows = []
        for row in sheet.iter_rows(values_only=True):
            cells = [str(cell) if cell is not None else "" for cell in row]
            if any(cells):
                rows.append(", ".join(cells))
        if rows:
            parts.append("\n".join(rows))
    return "\n\n".join(parts)[:_EXTRACTED_TEXT_LIMIT]


def _extract_csv(path: Path) -> str:
    with open(path, newline="", encoding="utf-8", errors="replace") as f:
        reader = csv.reader(f)
        rows = [", ".join(row) for row in reader]
    return "\n".join(rows)[:_EXTRACTED_TEXT_LIMIT]


def _extract_txt(path: Path) -> str:
    with open(path, "r", encoding="utf-8", errors="replace") as f:
        return f.read(_EXTRACTED_TEXT_LIMIT)


def _extract_image(path: Path) -> str:
    try:
        from PIL import Image
        import pytesseract
    except ImportError:
        logger.warning(
            "Pillow/pytesseract is not installed; image OCR skipped."
        )
        return ""

    try:
        image = Image.open(path)
        if image.mode not in ("L", "RGB"):
            image = image.convert("RGB")
        return pytesseract.image_to_string(image)[:_EXTRACTED_TEXT_LIMIT]
    except Exception as exc:
        logger.warning("OCR failed for %s: %s", path, exc)
        return ""
